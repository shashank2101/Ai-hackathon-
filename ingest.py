"""
ingest.py — Production RAG Ingestion Layer
==========================================
Converts any supported file format → markdown files in ./markdown_out/
Structure-aware chunking ready for vector DB ingestion.

PDF extraction chain (in order of preference):
  1. pdfminer.six  — best for text-heavy research PDFs
  2. pypdf          — fallback
  3. MarkItDown     — final fallback
  If none yield text → file is flagged as [NEEDS_OCR] (scanned/image PDF)

All other formats go through MarkItDown + direct handlers.
"""

import os
import re
import json
import base64
import sqlite3
import hashlib
import logging
import mimetypes
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

from markitdown import MarkItDown

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

MARKDOWN_OUT_DIR      = Path("markdown_out")
SQLITE_DB_PATH        = Path("reference_data.sqlite")
CHUNK_SIZE            = 500
CHUNK_OVERLAP         = 50
EXCEL_ROW_LIMIT       = 2000
JSON_ARRAY_ROW_LIMIT  = 5000

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ingest")

# ─────────────────────────────────────────────────────────────────────────────
# Extension sets
# ─────────────────────────────────────────────────────────────────────────────

DOCUMENT_EXTS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".rtf", ".odt"}
SHEET_EXTS    = {".xlsx", ".xls", ".csv"}
WEB_EXTS      = {".html", ".htm", ".xml"}
TEXT_EXTS     = {".txt", ".md", ".rst"}
DATA_EXTS     = {".json", ".yaml", ".yml", ".toml"}
STYLE_EXTS    = {".css", ".scss", ".sass", ".less"}
IMAGE_EXTS    = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff", ".tif"}

CODE_EXTS = {
    ".py", ".pyw", ".js", ".mjs", ".jsx", ".ts", ".tsx",
    ".java", ".c", ".h", ".cpp", ".cc", ".cxx", ".hpp",
    ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".kts",
    ".r", ".sh", ".bash", ".zsh", ".fish",
    ".sql", ".graphql", ".gql",
    ".lua", ".scala", ".ex", ".exs", ".hs",
    ".tf", ".tfvars", ".dockerfile", ".makefile",
}

_CODE_LANG: dict[str, str] = {
    ".py": "python", ".pyw": "python",
    ".js": "javascript", ".mjs": "javascript", ".jsx": "jsx",
    ".ts": "typescript", ".tsx": "tsx",
    ".java": "java", ".c": "c", ".h": "c",
    ".cpp": "cpp", ".cc": "cpp", ".cxx": "cpp", ".hpp": "cpp",
    ".cs": "csharp", ".go": "go", ".rs": "rust", ".rb": "ruby",
    ".php": "php", ".swift": "swift", ".kt": "kotlin", ".kts": "kotlin",
    ".r": "r", ".sh": "bash", ".bash": "bash", ".zsh": "bash", ".fish": "bash",
    ".sql": "sql", ".graphql": "graphql", ".gql": "graphql",
    ".lua": "lua", ".scala": "scala", ".ex": "elixir", ".exs": "elixir",
    ".hs": "haskell", ".tf": "hcl", ".tfvars": "hcl",
    ".dockerfile": "dockerfile", ".makefile": "makefile",
    ".css": "css", ".scss": "scss", ".sass": "sass", ".less": "less",
    ".yaml": "yaml", ".yml": "yaml", ".toml": "toml",
}

ALL_SUPPORTED_EXTS = (
    DOCUMENT_EXTS | SHEET_EXTS | WEB_EXTS | TEXT_EXTS |
    DATA_EXTS | STYLE_EXTS | IMAGE_EXTS | CODE_EXTS
)

# ─────────────────────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Chunk:
    source_file : str
    chunk_index : int
    content     : str
    heading_path: list[str] = field(default_factory=list)
    chunk_type  : str = "text"
    metadata    : dict = field(default_factory=dict)

    def token_estimate(self) -> int:
        return int(len(self.content.split()) * 0.75)

    def to_dict(self) -> dict:
        return {
            "source_file" : self.source_file,
            "chunk_index" : self.chunk_index,
            "content"     : self.content,
            "heading_path": " > ".join(self.heading_path),
            "chunk_type"  : self.chunk_type,
            "metadata"    : self.metadata,
        }


# ─────────────────────────────────────────────────────────────────────────────
# PDF extraction — multi-library fallback chain
# ─────────────────────────────────────────────────────────────────────────────

def _is_binary_garbage(text: str) -> bool:
    """
    Returns True if the extracted text looks like raw PDF binary
    (lots of non-printable bytes, stream markers, etc.) rather than
    actual document content.
    """
    if not text or len(text.strip()) < 50:
        return True
    # Count printable ASCII characters
    printable = sum(1 for c in text if c.isprintable() or c in '\n\r\t')
    ratio = printable / len(text)
    if ratio < 0.85:
        return True
    # Check for telltale PDF binary markers
    binary_markers = ["%PDF-", "endstream", "endobj", "xref", "startxref",
                      "FlateDecode", "/Filter", "/Length", "stream\n"]
    hits = sum(1 for m in binary_markers if m in text)
    if hits >= 3:
        return True
    return False


def extract_pdf_text(src: Path) -> tuple[str, str]:
    """
    Extract text from a PDF using ONLY MarkItDown.
    """
    # ── 1. MarkItDown ────────────────────────────────────────────────────────
    try:
        from markitdown import MarkItDown
        _md_conv = MarkItDown()
        result = _md_conv.convert(str(src))
        text = result.text_content or ""
        
        if text and not _is_binary_garbage(text) and len(text.strip()) > 50:
            log.info(f"  PDF extracted via MarkItDown ({len(text.split())} words)")
            return text, "markitdown"
            
        log.info("  MarkItDown: empty or binary output")
    except Exception as e:
        log.warning(f"  MarkItDown failed: {e}")

    # ── 2. Fallback — scanned/image PDF ──────────────────────────────────────
    log.warning(f"  {src.name}: MarkItDown failed — likely scanned/image PDF")
    return (
        f"# {src.name}\n\n"
        f"> **[NEEDS_OCR]** This PDF appears to be scanned or image-based. "
        f"No text could be extracted. Set `ANTHROPIC_API_KEY` and re-run with "
        f"`--ocr` flag to describe pages via vision LLM.\n"
    ), "needs_ocr"
def _pdf_text_to_markdown(raw_text: str, filename: str) -> str:
    """
    Clean up raw PDF-extracted text into decent markdown.
    - Remove excessive whitespace/blank lines
    - Try to detect headings (ALL CAPS short lines, numbered sections)
    - Preserve paragraph structure
    """
    lines = raw_text.splitlines()
    out: list[str] = [f"# {filename}\n"]
    prev_blank = False

    for line in lines:
        stripped = line.strip()

        # Skip blank lines but collapse multiples
        if not stripped:
            if not prev_blank:
                out.append("")
            prev_blank = True
            continue
        prev_blank = False

        # Heuristic: numbered section heading (e.g. "1 Introduction", "2.3 Methods")
        if re.match(r'^\d+(\.\d+)*\s+[A-Z]', stripped) and len(stripped) < 80:
            depth = stripped.split()[0].count('.') + 1
            hashes = '#' * min(depth + 1, 4)
            out.append(f"\n{hashes} {stripped}\n")
            continue

        # Heuristic: ALL CAPS short line → heading
        if stripped.isupper() and 3 < len(stripped) < 60 and ' ' in stripped:
            out.append(f"\n## {stripped.title()}\n")
            continue

        # Abstract / References / Conclusion / etc.
        if re.match(r'^(Abstract|References|Introduction|Conclusion|Related Work|Acknowledgements?)$',
                    stripped, re.IGNORECASE):
            out.append(f"\n## {stripped}\n")
            continue

        out.append(stripped)

    return "\n".join(out)


# ─────────────────────────────────────────────────────────────────────────────
# Image → LLM description
# ─────────────────────────────────────────────────────────────────────────────

def describe_image(src: Path) -> str:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        log.warning(f"  Image {src.name}: ANTHROPIC_API_KEY not set — placeholder used")
        return (
            f"**[Image: {src.name}]**\n\n"
            f"> *Description pending — set ANTHROPIC_API_KEY to enable LLM-based image description.*\n"
        )
    try:
        import anthropic
        mime_map = {
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png", ".gif": "image/gif", ".webp": "image/webp",
        }
        media_type = mime_map.get(src.suffix.lower(), "image/jpeg")
        img_data   = base64.standard_b64encode(src.read_bytes()).decode("utf-8")
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=512,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_data}},
                    {"type": "text", "text": (
                        "Describe this image in detail for a RAG knowledge base. "
                        "Include: what is shown, any text visible, layout/structure if it's a diagram or chart, "
                        "and any data or key information present. Be specific and factual. "
                        "Write 2-5 sentences in plain prose."
                    )},
                ],
            }],
        )
        description = response.content[0].text.strip()
        log.info(f"  Image {src.name}: described via LLM ({len(description.split())} words)")
        return f"**[Image: {src.name}]**\n\n{description}\n"
    except Exception as e:
        log.warning(f"  Image {src.name}: LLM description failed — {e}")
        return f"**[Image: {src.name}]**\n\n> *Image description failed: {e}*\n"


# ─────────────────────────────────────────────────────────────────────────────
# JSON type detection
# ─────────────────────────────────────────────────────────────────────────────

def _detect_json_type(data) -> str:
    if isinstance(data, list):
        if not data:
            return "mixed"
        if all(isinstance(r, dict) for r in data):
            sample   = data[:50]
            key_sets = [frozenset(r.keys()) for r in sample]
            if len(set(key_sets)) <= 3:
                return "records"
        return "mixed"
    if isinstance(data, dict):
        return "config"
    return "mixed"


def _flatten_dict(d: dict, prefix: str = "", sep: str = ".") -> dict:
    out = {}
    for k, v in d.items():
        key = f"{prefix}{sep}{k}" if prefix else k
        if isinstance(v, dict):
            out.update(_flatten_dict(v, key, sep))
        elif isinstance(v, list):
            out[key] = json.dumps(v, ensure_ascii=False)
        else:
            out[key] = v
    return out


def json_to_markdown(src: Path) -> tuple[str, str]:
    try:
        raw  = src.read_text(encoding="utf-8", errors="replace")
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        log.warning(f"  {src.name}: invalid JSON — {e}")
        return f"# {src.name}\n\n```json\n{src.read_text(errors='replace')}\n```\n", "mixed"

    jtype = _detect_json_type(data)

    if jtype == "records":
        records  = data
        all_keys = list(dict.fromkeys(k for r in records for k in r.keys()))
        header   = "| " + " | ".join(str(k) for k in all_keys) + " |"
        divider  = "| " + " | ".join("---" for _ in all_keys) + " |"
        rows     = []
        for r in records[:100]:
            cells = [str(r.get(k, "")).replace("|", "\\|").replace("\n", " ") for k in all_keys]
            rows.append("| " + " | ".join(cells) + " |")
        table = "\n".join([header, divider] + rows)
        note  = f"\n\n> *Preview: first 100 of {len(records)} records.*\n" if len(records) > 100 else ""
        return f"# {src.name}\n\n**Type:** JSON records — {len(records)} rows, {len(all_keys)} fields\n\n{table}{note}\n", "records"

    if jtype == "config":
        flat  = _flatten_dict(data)
        lines = [f"- **`{k}`**: `{v}`" for k, v in flat.items()]
        return f"# {src.name}\n\n**Type:** JSON configuration\n\n" + "\n".join(lines) + "\n", "config"

    pretty = json.dumps(data, indent=2, ensure_ascii=False)
    return f"# {src.name}\n\n```json\n{pretty}\n```\n", "mixed"


def json_records_to_sqlite(src: Path, data: list[dict], db_path: Path = SQLITE_DB_PATH):
    if len(data) > JSON_ARRAY_ROW_LIMIT:
        log.info(f"  {src.name}: {len(data)} records > limit — skipping SQLite export")
        return None
    all_keys  = list(dict.fromkeys(k for r in data for k in r.keys()))
    safe_cols = [re.sub(r'\W+', '_', str(k)).strip('_') or f"col_{i}" for i, k in enumerate(all_keys)]
    seen: dict[str, int] = {}
    deduped = []
    for c in safe_cols:
        if c in seen:
            seen[c] += 1
            deduped.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            deduped.append(c)
    table_name = re.sub(r'\W+', '_', src.stem).strip('_')
    col_defs   = ", ".join(f'"{c}" TEXT' for c in deduped)
    ph         = ", ".join("?" * len(deduped))
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
        for r in data:
            vals = []
            for k in all_keys:
                v = r.get(k)
                vals.append(json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else (str(v) if v is not None else None))
            conn.execute(f'INSERT INTO "{table_name}" VALUES ({ph})', vals)
        conn.commit()
        log.info(f"  SQLite: wrote {len(data)} rows → table '{table_name}'")
        return table_name
    except Exception as e:
        log.warning(f"  SQLite insert failed for {src.name}: {e}")
        return None
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Excel → SQLite
# ─────────────────────────────────────────────────────────────────────────────

def excel_sheets_to_sqlite(src: Path, db_path: Path = SQLITE_DB_PATH) -> list[str]:
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping SQLite export for Excel")
        return []
    written = []
    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"openpyxl cannot open {src.name}: {e}")
        return []
    conn = sqlite3.connect(db_path)
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = rows[0]
        if any(h is None for h in header_row):
            continue
        if len(rows) - 1 > EXCEL_ROW_LIMIT:
            continue
        cols = [re.sub(r'\W+', '_', str(h)).strip('_') or f"col_{i}" for i, h in enumerate(header_row)]
        seen: dict[str, int] = {}
        safe_cols = []
        for c in cols:
            if c in seen:
                seen[c] += 1
                safe_cols.append(f"{c}_{seen[c]}")
            else:
                seen[c] = 0
                safe_cols.append(c)
        table_name = re.sub(r'\W+', '_', f"{src.stem}__{sheet_name}").strip('_')
        col_defs   = ", ".join(f'"{c}" TEXT' for c in safe_cols)
        ph         = ", ".join("?" * len(safe_cols))
        try:
            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
            for row in rows[1:]:
                vals = [str(v) if v is not None else None for v in row[:len(safe_cols)]]
                conn.execute(f'INSERT INTO "{table_name}" VALUES ({ph})', vals)
            conn.commit()
            log.info(f"  SQLite: wrote {len(rows)-1} rows → '{table_name}'")
            written.append(table_name)
        except Exception as e:
            log.warning(f"  SQLite insert failed for sheet '{sheet_name}': {e}")
    conn.close()
    wb.close()
    return written


# ─────────────────────────────────────────────────────────────────────────────
# Main per-file converter
# ─────────────────────────────────────────────────────────────────────────────

_md_converter = MarkItDown()


def file_to_markdown(src: Path) -> tuple[str, str]:
    """Returns (markdown_text, file_category)."""
    suffix = src.suffix.lower()

    # Images
    if suffix in IMAGE_EXTS:
        return f"# Image: {src.name}\n\n{describe_image(src)}\n", "image"

    # PDF — use our own extraction chain, NOT MarkItDown directly
    if suffix == ".pdf":
        md_text, method = extract_pdf_text(src)
        return md_text, f"document_pdf_{method}"

    # Code
    if suffix in CODE_EXTS:
        lang = _CODE_LANG.get(suffix, "")
        try:
            text = src.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return f"# {src.name}\n\n[Unreadable]\n", "code"
        return f"# Code: {src.name}\n\n**Language:** {lang or suffix.lstrip('.')}\n\n```{lang}\n{text}\n```\n", "code"

    # Style files
    if suffix in STYLE_EXTS:
        lang = _CODE_LANG.get(suffix, "css")
        try:
            text = src.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return f"# {src.name}\n\n[Unreadable]\n", "style"
        return f"# Style: {src.name}\n\n**Type:** {lang.upper()}\n\n```{lang}\n{text}\n```\n", "style"

    # JSON
    if suffix == ".json":
        md, jtype = json_to_markdown(src)
        return md, f"json_{jtype}"

    # YAML / TOML
    if suffix in (".yaml", ".yml", ".toml"):
        lang = _CODE_LANG.get(suffix, "yaml")
        try:
            text = src.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return f"# {src.name}\n\n[Unreadable]\n", "data"
        try:
            result = _md_converter.convert(str(src))
            md_text = result.text_content or ""
            if md_text.strip():
                return f"# Config: {src.name}\n\n**Format:** {lang.upper()}\n\n{md_text}\n\n```{lang}\n{text}\n```\n", "data"
        except Exception:
            pass
        return f"# Config: {src.name}\n\n**Format:** {lang.upper()}\n\n```{lang}\n{text}\n```\n", "data"

    # Everything else via MarkItDown
    try:
        result = _md_converter.convert(str(src))
        md = result.text_content or ""
        if md.strip() and not _is_binary_garbage(md):
            category = (
                "sheet"    if suffix in SHEET_EXTS    else
                "document" if suffix in DOCUMENT_EXTS else
                "web"      if suffix in WEB_EXTS      else
                "text"
            )
            return md, category
    except Exception as e:
        log.warning(f"MarkItDown failed for {src.name}: {e}")

    # Last resort plain text
    try:
        text = src.read_text(errors="replace")
        if not _is_binary_garbage(text):
            return f"# {src.name}\n\n" + text, "text"
    except Exception:
        pass

    return f"# {src.name}\n\n[Binary or unreadable — skipped]\n", "text"


# ─────────────────────────────────────────────────────────────────────────────
# Structure-aware chunker
# ─────────────────────────────────────────────────────────────────────────────

_HEADING_RE  = re.compile(r'^(#{1,6})\s+(.+)$')
_FENCE_START = re.compile(r'^```')
_TABLE_ROW   = re.compile(r'^\|')


def _word_count(text: str) -> int:
    return len(text.split())


def _split_text_block(text: str, chunk_size: int, overlap: int) -> list[str]:
    sentences  = re.split(r'(?<=[.!?])\s+', text)
    sub_chunks: list[str] = []
    current: list[str]    = []
    current_wc = 0
    for sent in sentences:
        wc = _word_count(sent)
        if current_wc + wc > chunk_size and current:
            sub_chunks.append(" ".join(current))
            overlap_words = " ".join(current)[-overlap * 6:].split()[-overlap:]
            current   = overlap_words + [sent]
            current_wc = _word_count(" ".join(current))
        else:
            current.append(sent)
            current_wc += wc
    if current:
        sub_chunks.append(" ".join(current))
    return sub_chunks


def chunk_markdown(md_text: str, source_file: str,
                   chunk_size: int = CHUNK_SIZE,
                   overlap: int = CHUNK_OVERLAP,
                   file_category: str = "text") -> list[Chunk]:

    # Code, style, image → single chunk (never split)
    if file_category in ("code", "style", "image"):
        return [Chunk(
            source_file=source_file, chunk_index=0,
            content=md_text.strip(),
            heading_path=[], chunk_type=file_category,
            metadata={"file_category": file_category},
        )]

    chunks: list[Chunk]   = []
    heading_stack: list[str] = []
    lines       = md_text.splitlines(keepends=True)
    buffer: list[str] = []
    buffer_type = "text"
    in_fence    = False
    fence_buffer: list[str] = []
    in_table    = False
    table_buffer: list[str] = []
    idx = 0

    def flush(buf: list[str], btype: str):
        nonlocal idx
        text = "".join(buf).strip()
        if not text:
            return
        if btype == "text" and _word_count(text) > chunk_size:
            for sub in _split_text_block(text, chunk_size, overlap):
                chunks.append(Chunk(source_file=source_file, chunk_index=idx,
                                    content=sub, heading_path=list(heading_stack),
                                    chunk_type=btype, metadata={"file_category": file_category}))
                idx += 1
        else:
            chunks.append(Chunk(source_file=source_file, chunk_index=idx,
                                content=text, heading_path=list(heading_stack),
                                chunk_type=btype, metadata={"file_category": file_category}))
            idx += 1

    for line in lines:
        stripped = line.rstrip('\n')
        if _FENCE_START.match(stripped):
            if not in_fence:
                flush(buffer, buffer_type); buffer = []
                in_fence = True; fence_buffer = [line]
            else:
                fence_buffer.append(line)
                flush(fence_buffer, "code"); fence_buffer = []; in_fence = False
            continue
        if in_fence:
            fence_buffer.append(line); continue
        if _TABLE_ROW.match(stripped):
            if not in_table:
                flush(buffer, buffer_type); buffer = []
                in_table = True; table_buffer = [line]
            else:
                table_buffer.append(line)
            continue
        else:
            if in_table:
                flush(table_buffer, "table"); table_buffer = []; in_table = False
        m = _HEADING_RE.match(stripped)
        if m:
            level = len(m.group(1)); title = m.group(2).strip()
            flush(buffer, buffer_type); buffer = []
            heading_stack = heading_stack[:level - 1]; heading_stack.append(title)
            buffer.append(line); buffer_type = "text"; continue
        buffer.append(line)
        if stripped == "" and _word_count("".join(buffer)) >= chunk_size:
            flush(buffer, buffer_type)
            tail_words = " ".join("".join(buffer).split()[-overlap:]) if overlap else ""
            buffer = [tail_words + "\n"] if tail_words else []; buffer_type = "text"

    flush(fence_buffer if in_fence else (table_buffer if in_table else buffer),
          "code" if in_fence else ("table" if in_table else buffer_type))
    return chunks


# ─────────────────────────────────────────────────────────────────────────────
# Main ingestion pipeline
# ─────────────────────────────────────────────────────────────────────────────

def ingest_file(src: Path, out_dir: Path = MARKDOWN_OUT_DIR) -> list[Chunk]:
    suffix = src.suffix.lower()
    if src.name.lower() in ("dockerfile", "makefile", "gemfile", "procfile"):
        suffix = f".{src.name.lower()}"
    if suffix not in ALL_SUPPORTED_EXTS:
        log.warning(f"Skipping unsupported: {src.name}")
        return []

    log.info(f"Processing: {src.name}")

    md_text, file_category = file_to_markdown(src)

    file_hash = hashlib.md5(src.read_bytes()).hexdigest()[:8]
    meta_header = (
        f"<!-- source: {src.name} | category: {file_category} | "
        f"ingested: {datetime.utcnow().isoformat(timespec='seconds')}Z | "
        f"hash: {file_hash} -->\n\n"
    )
    md_text = meta_header + md_text

    # Excel → SQLite
    if suffix in (".xlsx", ".xls"):
        tables = excel_sheets_to_sqlite(src)
        if tables:
            md_text += f"\n\n---\n> **SQLite tables written:** {', '.join(tables)}\n"

    # JSON records → SQLite
    if suffix == ".json" and file_category == "json_records":
        try:
            data  = json.loads(src.read_text(encoding="utf-8", errors="replace"))
            table = json_records_to_sqlite(src, data)
            if table:
                md_text += f"\n\n---\n> **SQLite table written:** {table}\n"
        except Exception as e:
            log.warning(f"  JSON SQLite export failed: {e}")

    # Write markdown
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_stem   = re.sub(r'\W+', '_', src.stem)
    md_path     = out_dir / f"{safe_stem}__{file_hash}.md"
    md_path.write_text(md_text, encoding="utf-8")
    log.info(f"  → {md_path}  [{file_category}]")

    # Chunk
    chunks  = chunk_markdown(md_text, source_file=str(src), file_category=file_category)
    avg_tok = sum(c.token_estimate() for c in chunks) // max(len(chunks), 1)
    log.info(f"  → {len(chunks)} chunks  avg≈{avg_tok} tokens")
    return chunks


def ingest_directory(src_dir: Path, out_dir: Path = MARKDOWN_OUT_DIR,
                     recursive: bool = True) -> list[Chunk]:
    glob_fn    = src_dir.rglob if recursive else src_dir.glob
    all_chunks: list[Chunk] = []
    for f in sorted(glob_fn("*")):
        if f.is_file() and f.suffix.lower() in ALL_SUPPORTED_EXTS:
            all_chunks.extend(ingest_file(f, out_dir))
    return all_chunks


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse, sys
    from collections import Counter

    parser = argparse.ArgumentParser(description="RAG Ingestion Layer")
    parser.add_argument("sources", nargs="+")
    parser.add_argument("--out",         default="markdown_out")
    parser.add_argument("--db",          default="reference_data.sqlite")
    parser.add_argument("--chunk-size",  type=int, default=500)
    parser.add_argument("--overlap",     type=int, default=50)
    parser.add_argument("--no-recursive", action="store_true")
    args = parser.parse_args()

    MARKDOWN_OUT_DIR = Path(args.out)
    SQLITE_DB_PATH   = Path(args.db)
    CHUNK_SIZE       = args.chunk_size
    CHUNK_OVERLAP    = args.overlap

    total_chunks: list[Chunk] = []
    for s in args.sources:
        p = Path(s)
        if p.is_dir():
            total_chunks.extend(ingest_directory(p, MARKDOWN_OUT_DIR, not args.no_recursive))
        elif p.is_file():
            total_chunks.extend(ingest_file(p, MARKDOWN_OUT_DIR))
        else:
            log.error(f"Not found: {s}"); sys.exit(1)

    by_type = Counter(c.chunk_type for c in total_chunks)
    print(f"\n{'─'*50}")
    print(f"  Total chunks : {len(total_chunks)}")
    for ctype, n in sorted(by_type.items()):
        print(f"  {ctype:<14}: {n}")
    print(f"  Markdown out : {MARKDOWN_OUT_DIR}/")
    if SQLITE_DB_PATH.exists():
        print(f"  SQLite DB    : {SQLITE_DB_PATH}")
    print(f"{'─'*50}")